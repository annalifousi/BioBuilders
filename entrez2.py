import os
from Bio import Entrez
import openpyxl
from openpyxl import Workbook
import csv
from bs4 import BeautifulSoup

# Identifying the connected user
Entrez.email = 'annalifousi@gmail.com'
Entrez.tool = 'Demoscript'

def search_entrez_and_save(query, xml_path, excel_path):
    # Searching for the query in Entrez
    info = Entrez.esearch(db="pubmed", retmax=100000, term=query)
    # Parsing the XML data
    record = Entrez.read(info)

    # Count the number of articles
    num_articles = len(record['IdList'])

    # Retrieve records in XML format
    fetch = Entrez.efetch(db='pubmed',
                          retmode='xml',
                          id=record['IdList'],
                          rettype='full')

    # Write records in XML file
    with open(xml_path, 'wb') as f:
        f.write(fetch.read())

    fetch.close()  # Close the fetch object to release resources

    # Print the number of articles and the file path
    print(f'Number of articles returned: {num_articles}')
    print(f'The file has been saved to: {xml_path}')

    # Create a new Excel workbook if it doesn't exist, otherwise load it
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"
        ws.append(["Query", "Number of Articles"])
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

    # Check the last entry to avoid duplication
    last_row = ws.max_row
    if last_row > 1:  # Ensure there's more than just the header
        last_query = ws.cell(row=last_row, column=1).value
        last_num_articles = ws.cell(row=last_row, column=2).value
        if last_query == query and last_num_articles == num_articles:
            print("The query and number of articles are the same as the last entry. No new entry added.")
        else:
            ws.append([query, num_articles])
            print(f'New entry added: {query} - {num_articles}')
    else:
        ws.append([query, num_articles])
        print(f'New entry added: {query} - {num_articles}')

    # Save the workbook
    wb.save(excel_path)

    print(f'The Excel file has been updated and saved to: {excel_path}')

def parse_xml_to_csv(xml_path, csv_path):
    # Open and read the XML file
    with open(xml_path, 'r', encoding='utf-8') as file:
        xml_content = file.read()

    # Parse the XML content with BeautifulSoup
    soup = BeautifulSoup(xml_content, 'xml')

    # Open the CSV file for writing
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        # Write the header
        csvwriter.writerow(['PMID', 'Authors', 'Title', 'Date', 'Abstract', 'PMC ID'])

        # Find all articles in the XML
        articles = soup.find_all('PubmedArticle')
        pmc_count = 0

        for i, article in enumerate(articles):
            # Extract PMID
            pmid = article.find('PMID').text if article.find('PMID') else 'N/A'

            # Extract authors
            authors_list = []
            for author in article.find_all('Author'):
                last_name = author.find('LastName')
                first_name = author.find('ForeName')
                if last_name and first_name:
                    authors_list.append(f"{first_name.text} {last_name.text}")
            authors = ', '.join(authors_list)

            # Extract title
            title = article.find('ArticleTitle')
            title_text = title.text if title else 'N/A'

            # Extract date
            pub_date = article.find('PubDate')
            if pub_date:
                year = pub_date.find('Year')
                medline_date = pub_date.find('MedlineDate')
                date_text = year.text if year else medline_date.text if medline_date else 'N/A'
            else:
                date_text = 'N/A'

            # Extract abstract
            abstract = article.find('Abstract')
            if abstract:
                abstract_text = ' '.join([p.text for p in abstract.find_all('AbstractText')])
            else:
                abstract_text = 'N/A'

            # Extract PMC ID if available
            pmc_id = None
            for article_id in article.find_all('ArticleId'):
                if article_id.get('IdType') == 'pmc':
                    pmc_id = article_id.text
                    pmc_count += 1
                    break

            # Debug information
            print(f"Article {i+1}: PMID={pmid}, PMC ID={pmc_id}")

            # Write the data to CSV only if PMC ID is available
            if pmc_id:
                csvwriter.writerow([pmid, authors, title_text, date_text, abstract_text, pmc_id])

    print(f'The CSV file has been created and saved to: {csv_path}')
    print(f'Number of articles with available PMC ID: {pmc_count}')
    return pmc_count


# Entrez searching/ getting xml file with articles
query = (
    '("Endocrine Disruptors"[MeSH Terms] OR "endocrine disrupting chemicals"[Title/Abstract] OR "EDCs"[Title/Abstract] OR "hormonally active agents"[Title/Abstract] OR "Endocrine disrupting compounds"[Title/Abstract]) '
    'AND (human[Title/Abstract]) '
    'AND ("Receptors, Endocrine"[MeSH Terms] OR "receptors"[Title/Abstract] OR "receptor"[Title/Abstract] OR target[Title/Abstract]) '
    'AND ('
        '("human estrogen receptor alpha"[Title/Abstract] OR "HERa"[Title/Abstract] OR "Estrogen"[Title/Abstract] OR target[Title/Abstract]) '
        'OR "estradiol"[Title/Abstract] '
        'OR ("HERb"[Title/Abstract] OR "Human estrogen receptor beta"[Title/Abstract]) '
        'OR ("binding"[Title/Abstract] AND ("Human androgen receptor"[Title/Abstract] OR "HAR"[Title/Abstract] OR "Testosterone"[Title/Abstract])) '
        'OR ("NR3C1"[Title/Abstract] OR "Human glucocorticoid receptor"[Title/Abstract]) '
        'OR "Dexamethasone"[Title/Abstract] '
        'OR ("Human mineralocorticoid receptor"[Title/Abstract] OR "Aldosterone"[Title/Abstract] OR "NR3C2"[Title/Abstract]) '
        'OR ("Progesterone"[Title/Abstract] OR "NR3C3"[Title/Abstract] OR "Human progesterone receptor"[Title/Abstract]) '
    ') '
    'OR ('
        '("Endocrine Disrupting Chemicals"[MeSH Terms] OR "Endocrine Disruptors"[MeSH Terms] OR "EDCs"[Title/Abstract] OR "Endocrine Disruptors"[Title/Abstract]) '
        'AND ('
            '("Bisphenol A"[MeSH Terms] OR "BPA"[Title/Abstract]) '
            'OR ("Phthalates"[MeSH Terms] OR "Phthalate"[Title/Abstract]) '
            'OR ("Polychlorinated Biphenyls"[MeSH Terms] OR "PCBs"[Title/Abstract]) '
           # 'OR ("Regulation"[MeSH Terms] OR "Risk Assessment"[Title/Abstract]) '
        ') '
    ') '
    'NOT review[Publication Type]'
)
xml_path = '../BioBuilders/articles_tool.xml'
excel_path = '../BioBuilders/query_results.xlsx'

search_entrez_and_save(query, xml_path, excel_path)

# xml scrapping/ getting csv with info
xml_path = '../BioBuilders/articles_tool.xml'
csv_path = '../BioBuilders/articles.csv'

pmc_count = parse_xml_to_csv(xml_path, csv_path)