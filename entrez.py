import os
from Bio import Entrez
import openpyxl
from openpyxl import Workbook
import csv
from bs4 import BeautifulSoup
import pandas as pd

# Identifying the connected user
Entrez.email = 'annalifousi@gmail.com'
Entrez.tool = 'Demoscript'



# Data retrieval. We convert the data retrieved from the pubmed in xml format and then load them in an xml file (articles_toool.xml)


def search_entrez_and_save(query, xml_path, excel_path):
    # Searching for the query in Entrez
    info = Entrez.esearch(db="pubmed", retmax=100000, term=query)
    # Parsing the XML data
    record = Entrez.read(info)

    # Count the number of articles
    num_articles = len(record['IdList'])
    
    #if we don't find articles
    if num_articles == 0:
        print("No articles found.")
        return

    # Retrieve records in XML format
    fetch = Entrez.efetch(db='pubmed',
                          retmode='xml',
                          id=record['IdList'],
                          rettype='full')

    # Write records in XML file
    with open(xml_path, 'wb') as f:
        f.write(fetch.read())

    fetch.close()  # Close the fetch object to release resources

    # Print the number of articles and the file path
    print(f'Number of articles returned: {num_articles}')
    print(f'The file has been saved to: {xml_path}')

    # Create a new Excel workbook if it doesn't exist, otherwise load it
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"
        ws.append(["Query", "Number of Articles"])
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

    # Check the last entry to avoid duplication
    last_row = ws.max_row
    if last_row > 1:  # Ensure there's more than just the header
        last_query = ws.cell(row=last_row, column=1).value
        last_num_articles = ws.cell(row=last_row, column=2).value
        if last_query == query and last_num_articles == num_articles:
            print("The query and number of articles are the same as the last entry. No new entry added.")
        else:
            ws.append([query, num_articles])
            print(f'New entry added: {query} - {num_articles}')
    else:
        ws.append([query, num_articles])
        print(f'New entry added: {query} - {num_articles}')

    # Save the workbook
    wb.save(excel_path)

    print(f'The Excel file has been updated and saved to: {excel_path}')

def parse_xml_to_csv(xml_path, csv_path):
    # Open and read the XML file
    with open(xml_path, 'r', encoding='utf-8') as file:
        xml_content = file.read()


#Web scraping procedure 

    # Parse the XML content with BeautifulSoup
    soup = BeautifulSoup(xml_content, 'xml')

    # Open the CSV file for writing
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        # Write the header
        csvwriter.writerow(['PMID', 'Authors', 'Title', 'Date', 'Abstract', 'PMC ID'])

        # Find all articles in the XML
        articles = soup.find_all('PubmedArticle')
        pmc_count = 0

        for i, article in enumerate(articles):
            # Extract PMID
            pmid = article.find('PMID').text if article.find('PMID') else 'N/A'

            # Extract authors
            authors_list = []
            for author in article.find_all('Author'):
                last_name = author.find('LastName')
                first_name = author.find('ForeName')
                if last_name and first_name:
                    authors_list.append(f"{first_name.text} {last_name.text}")
            authors = ', '.join(authors_list)

            # Extract title
            title = article.find('ArticleTitle')
            title_text = title.text if title else 'N/A'

            # Extract date
            pub_date = article.find('PubDate')
            if pub_date:
                year = pub_date.find('Year')
                medline_date = pub_date.find('MedlineDate')
                date_text = year.text if year else medline_date.text if medline_date else 'N/A'
            else:
                date_text = 'N/A'

            # Extract abstract
            abstract = article.find('Abstract')
            if abstract:
                abstract_text = ' '.join([p.text for p in abstract.find_all('AbstractText')])
            else:
                abstract_text = 'N/A'
                
            # Extract PMC ID if available
            pmc_id = None
            for article_id in article.find_all('ArticleId'):
                if article_id.get('IdType') == 'pmc':
                    pmc_id = article_id.text
                    pmc_count += 1
                    break

            # Debug information
            print(f"Article {i+1}: PMID={pmid}, PMC ID={pmc_id}")

            # Write the data to CSV only if PMC ID is available
            if pmc_id:
                csvwriter.writerow([pmid, authors, title_text, date_text, abstract_text, pmc_id])

    print(f'The CSV file has been created and saved to: {csv_path}')
    print(f'Number of articles with available PMC ID: {pmc_count}')
    return pmc_count



# Load the TSV files for the scrapping

# Read EDCs from androgen_EKDB.tsv file
print("Reading EDCs from androgen_EKDB.tsv...")
androgen_df = pd.read_csv('BioBuilders/EDC_androgen_catalog.tsv', sep='\t')
edc_androgen_names = set(androgen_df['Name'].str.lower().unique())  # Get unique EDC names and convert to lowercase

# Read EDCs from estrogen.tsv file from the EDKB database
print("Reading EDCs from estrogen.tsv...")
estrogen_df = pd.read_csv('BioBuilders/EDC_estrogen_catalog.tsv', sep='\t')
edc_estrogen_names = set(estrogen_df['Name'].str.lower().unique())  # Get unique EDC names and convert to lowercase

# Read EDCs from EDC_catalog_deduct.tsv file from the deduct database
print("Reading EDCs from EDC_catalog_deduct.tsv...")
deduct_df = pd.read_csv('BioBuilders/EDC_catalog_deduct.tsv', sep='\t')
edc_deduct_names = set(deduct_df['Name'].str.lower().unique())  # Get unique EDC names and convert to lowercase

# Read EDCs from manual_catalog.tsv file
print("Reading EDCs from manual_catalog.tsv...")
manual_df = pd.read_csv('BioBuilders/manual_catalog.tsv', sep='\t')
manual_df_names = set(manual_df['Name'].str.lower())  # Get unique EDC names and convert to lowercase

# Combine all unique EDC names
all_edc_names = edc_androgen_names.union(edc_estrogen_names).union(edc_deduct_names).union(manual_df_names)

# Create a DataFrame with the specified column name
all_edc_names_df = pd.DataFrame(list(all_edc_names), columns=['Name'])

# Save the DataFrame to a TSV file
all_edc_names_df.to_csv('BioBuilders/combined_edc_catalog.tsv', sep='\t', index=False)

# Read the combined TSV file into a DataFrame
tsv_file_path = 'BioBuilders/combined_edc_catalog.tsv'
combined_edc_df = pd.read_csv(tsv_file_path, sep='\t')

edc_keywords = combined_edc_df['Name'].tolist()

# Add the receptor keywords in the pipeline
receptors_file_path = 'Biobuilders/receptors.tsv'
receptors_df = pd.read_csv(receptors_file_path, sep='\t')
receptors_keywords = receptors_df['Name'].tolist()

#Add the activity query in the pipeline
action_file_path = 'Biobuilders/edc_vocabulary.tsv'
action_df = pd.read_csv(action_file_path, sep='\t')
action_keywords = action_df['Name'].tolist()


# Construct the keyword part of the query (EDCs)
edc_query = ' OR '.join([f'"{keyword}"[Title/Abstract]' for keyword in edc_keywords])

# Construct the keyword part of the query (receptors)
receptors_query = ' OR '.join([f'"{keyword}"[Title/Abstract]' for keyword in receptors_keywords])

# Construct the keyword part of the query (receptors)
action_query = ' OR '.join([f'"{keyword}"[Title/Abstract]' for keyword in action_keywords])


# MESH TERMS
# Construct the MESH terms part of the query (EDCs)
edc_mesh_query = ' OR '.join([f'"{term}"[MeSH Terms]' for term in edc_keywords])

# Construct the MESH terms part of the query (receptors)
receptors_mesh_query = ' OR '.join([f'"{term}"[MeSH Terms]' for term in receptors_keywords])

# Construct the MESH terms part of the query (action)
action_mesh_query = ' OR '.join([f'"{term}"[MeSH Terms]' for term in action_keywords])

#########################################################################
# Combine action-related terms with EDC-related terms using AND
edc_and_action_query = ' AND '.join([
    edc_query,
    edc_mesh_query,
    ' OR '.join([
        action_query,
        action_mesh_query
    ])
])



#########################################################################


#Additional query
# Entrez searching/ getting xml file with articles
query = (
    '("Endocrine Disruptors"[MeSH Terms] OR "endocrine disrupting chemicals"[Title/Abstract] OR "EDCs"[Title/Abstract] OR "hormonally active agents"[Title/Abstract] OR "Endocrine disrupting compounds"[Title/Abstract]) '
    #incorporated MIE search in the query
    'OR ("molecular initiating event"[Title/Abstract] OR "MIE"[Title/Abstract] OR "gene expression"[Title/Abstract])'
    'AND ('
        '("human estrogen receptor alpha"[Title/Abstract] OR "HERa"[Title/Abstract] OR target[Title/Abstract]) '
        'OR ("HERb"[Title/Abstract] OR "Human estrogen receptor beta"[Title/Abstract]) '
        'OR ("binding"[Title/Abstract] AND ("Human androgen receptor"[Title/Abstract] OR "HAR"[Title/Abstract])) '
        'OR ("NR3C1"[Title/Abstract] OR "Human glucocorticoid receptor"[Title/Abstract]) '
        'OR "NR3C2"[Title/Abstract]) '
        'OR "NR3C3"[Title/Abstract] OR "Human progesterone receptor"[Title/Abstract]) '
    ') '

        ') '
    ') '
    'NOT review[Publication Type]'

    #mesh terms
    'AND ("Reproduction"[MeSH Terms] OR "Endocrine System"[MeSH Terms] OR "chemistry"[MeSH Terms] OR "toxicity"[MeSH Terms])'
    'OR "Disruptors, Endocrine"[MeSH Terms] OR "Endocrine Disruptor"[MeSH Terms] OR "Endocrine Disrupting Chemical"[MeSH Terms] OR "Chemical, Endocrine Disrupting"[MeSH Terms] OR "Disrupting Chemical, Endocrine"[MeSH Terms] OR "Endocrine Disruptor Effect"[MeSH Terms] OR "Effect, Endocrine Disruptor"[MeSH Terms] OR "Endocrine Disruptor Effects"[MeSH Terms]'
    'OR "Water pollutants"[MeSH Terms] OR  "Gonadal Steroid Hormones"[MeSH Terms] OR "fertility"[MeSH Terms] OR "infertility"[MeSH Terms] OR "Androgen antagonists"[MeSH Terms] OR "astrogen antagonists"[MeSH Terms] OR "estrogen agonists"[MeSH Terms] OR "androgen agonists"[MeSH Terms] '
)

#combine everything

final_query = f'({query}) OR '.join([
    edc_and_action_query,
    receptors_query,
    receptors_mesh_query
])
xml_path = './articles_tool.xml'
excel_path = './query_results.xlsx'

#search_entrez_and_save(query1, xml_path, excel_path)
search_entrez_and_save(query, xml_path, excel_path)

# xml scrapping/ getting csv with info
xml_path = 'BioBuilders/articles_tool.xml'
csv_path = 'BioBuilders/articles.csv'

pmc_count = parse_xml_to_csv(xml_path, csv_path)